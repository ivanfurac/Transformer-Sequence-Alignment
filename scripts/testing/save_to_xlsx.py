import os
import openpyxl as xl

def get_info(file):

    with open(file, 'r') as f:
        lines = f.readlines()
    
    train_loss = []
    train_acc = []
    val_loss = []
    val_acc = []
    time = 0

    for line in lines:
        if 'Total params:' in line:
            parameters = int(line.strip('\n').split(' ')[2].replace(',', ''))
        if 'loss' in line:
            split = line.strip('\n').split(' ')
            train_loss.append(float(split[split.index('loss:') + 1]))
            train_acc.append(float(split[split.index('accuracy_function:') + 1]))
            val_loss.append(float(split[split.index('val_loss:') + 1]))
            val_acc.append(float(split[split.index('val_accuracy_function:') + 1]))
            time += int(split[split.index('val_accuracy_function:') + 3].split('s')[0])
            
    return min(train_loss), max(train_acc), min(val_loss), max(val_acc), time, parameters

path = 'C:/Users/Ivan/Desktop/CODE/models/EXCEL/'
excel_path = 'C:/Users/Ivan/Desktop/Diplomski Rad/Rezultati.xlsx'

wb = xl.load_workbook(excel_path)
ws = wb['Sheet1']

for file in os.listdir(path):
    row = int(file.split('.')[0])
    tloss, tacc, vloss, vacc, time, parameters = get_info(path + file)
    ws['D' + str(row)] = tacc
    ws['E' + str(row)] = tloss
    ws['F' + str(row)] = vacc
    ws['G' + str(row)] = vloss
    ws['H' + str(row)] = time
    ws['I' + str(row)] = parameters
    
wb.save(excel_path)